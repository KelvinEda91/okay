import openpyxl

# load workbook and select sheet
workbook = openpyxl.load_workbook('myspreadsheet.xlsx')
sheet = workbook.active

# read data from sheet
for row in sheet.iter_rows(min_row=2, values_only=True):
    food, car, country = row
    print(f"Name: {food}, Quantity: {car}, Price: {country}")


import openpyxl
worksheet = openpyxl.load_workbook("codespeedy.xlsx")
sheet = worksheet.active
for row in sheet.iter_rows(min_row=1, min_col=1, max_row=6, max_col=2):
    for cell in row:
        print(cell.value, end=" ")
    print()